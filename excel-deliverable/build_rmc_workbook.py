#!/usr/bin/env python3
"""Generate an advanced, connected RMC plant daily-operations workbook.

Env overrides (used for fast validation builds):
  RMC_TXN_ROWS     rows per transaction sheet   (default 6000)
  RMC_MASTER_ROWS  rows per setup/master sheet  (default 500)
  RMC_OUT          output path
"""

import os
import glob
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule, CellIsRule, FormulaRule

TXN_ROWS = int(os.environ.get("RMC_TXN_ROWS", "6000"))
MASTER_ROWS = int(os.environ.get("RMC_MASTER_ROWS", "500"))
OUT = os.environ.get("RMC_OUT", "excel-deliverable/RMC-Plant-Daily-Workbook.xlsx")

# ---------- palette (CONCRETE KING theme) ----------
NAVY = "08111F"; NAVY2 = "12203A"; GOLD = "F7C948"; GREEN = "22C55E"
BLUE = "38BDF8"; RED = "EF4444"; AMBER = "F59E0B"; WHITE = "FFFFFF"
INPUT_BG = "FFFDF3"; CALC_BG = "EEF6FF"

# ---------- company (printed on challan / batch report) ----------
COMPANY = "AAKRUTI INFRA"
COMPANY_SUB = "RMC Plant, Panvel  \u2022  +91 7498286760  \u2022  sales@aakrutiinfra.com"
PLANT_NAME = "AAKRUTI INFRA RMC PLANT"
PLANT_ADDR = ("Plant Add : Plot No. 77/2, JNPT Road, Behind Nilesh Dhaba, "
              "Karanjade, Tal-Panvel, Dist-Raigad 410 206")
OFFICE_ADDR = ("Office Add : Shop No. 45, Silver Crest Akshar, Plot 29, "
               "Sector 25, Navi Mumbai 410 206")
REG_LINE = ("GSTIN: 27ACKFA9816C1ZC      PAN: ACKFA9816C      "
            "Contact: 7498286760 / 9082189911      Mail: aakruti.infrarmc@gmail.com")
WEB_LINE = "www.aakrutiinfrarmc.com      qccomplaint@aakrutiinfrarmc.com      sales@aakrutiinfrarmc.com"
CEMENT_BRAND = "OPC 53 (Ultratech)"   # plant default, editable on the challan
ADMIX_TYPE = "PCE"                      # plant default, editable on the challan
DEFAULT_SLUMP = "180 mm"
DEFAULT_CUBE = "Yes"
CHALLAN_TERMS = [
    "\u2022 If the concrete is unloaded after 2.30 hours from batching, the client is fully responsible for the qty in the bill and for any TM damage.",
    "\u2022 Transit mixers must be released within ~1 hour. Retention beyond this attracts a penalty (around Rs 1200 per hour or part thereof).",
    "\u2022 Addition of water at site to regain workability is prohibited, or must be done only under strict technical supervision.",
    "\u2022 The supplier requires assurance of continuous pouring and may charge extra for small loads (e.g. less than 3-4 CuM).",
]

DATA_START = 5
HDR_ROW = 4
def end_row(rows): return DATA_START + rows - 1
MEND = end_row(MASTER_ROWS)   # master end
TEND = end_row(TXN_ROWS)      # transaction end

# shared style objects (reused -> small file, fast write)
thin = Side(style="thin", color="C7D0E0")
B = Border(left=thin, right=thin, top=thin, bottom=thin)
F_INPUT = Font(color=NAVY, size=11)
F_CALC = Font(color="1F3D6B", size=11)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_BG)
FILL_CALC = PatternFill("solid", fgColor=CALC_BG)
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_NAVY2 = PatternFill("solid", fgColor=NAVY2)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
H_FONT = Font(bold=True, color=WHITE, size=11)
TITLE_FONT = Font(bold=True, color=GOLD, size=20)
SUB_FONT = Font(color="9AA7BD", size=10, italic=True)
LABEL_FONT = Font(bold=True, color=NAVY, size=11)

wb = Workbook()


def style_header(ws, row, ncols, fill=FILL_NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H_FONT; cell.fill = fill; cell.border = B
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def title_block(ws, title, subtitle, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    t = ws.cell(row=1, column=1, value=title); t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    s = ws.cell(row=2, column=1, value=subtitle); s.font = SUB_FONT
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for r in (1, 2):
        for c in range(1, span + 1):
            ws.cell(row=r, column=c).fill = FILL_NAVY
    ws.row_dimensions[1].height = 34; ws.row_dimensions[2].height = 18


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_list_validation(ws, named_range, target_range):
    dv = DataValidation(type="list", formula1=f"={named_range}", allow_blank=True)
    dv.error = "Pick a value from the list."; dv.prompt = "Choose from the dropdown."
    ws.add_data_validation(dv); dv.add(target_range)


def inline_list(ws, items, target_range):
    dv = DataValidation(type="list", formula1='"' + ",".join(items) + '"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(target_range)


def defname(name, ref):
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def build_sheet(name, title, subtitle, columns, rows, examples=None, accent=FILL_NAVY):
    """columns: list of {h, w, type, formula(optional with {r})}"""
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ncols = len(columns)
    set_widths(ws, [c["w"] for c in columns])
    title_block(ws, title, subtitle, ncols)
    for i, c in enumerate(columns, start=1):
        ws.cell(row=HDR_ROW, column=i, value=c["h"])
    style_header(ws, HDR_ROW, ncols, fill=accent)
    ws.freeze_panes = f"A{DATA_START}"
    if examples:
        for ri, row in enumerate(examples, start=DATA_START):
            for ci, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=ri, column=ci, value=val)
    end = end_row(rows)
    fmtmap = {"date": "dd-mmm-yyyy", "num": "#,##0.00", "money": '\u20b9#,##0.00',
              "int": "#,##0", "pct": "0.00", "time": "h:mm AM/PM"}
    for i, c in enumerate(columns, start=1):
        L = get_column_letter(i)
        t = c.get("type", "input")
        is_calc = t == "calc"
        fill = FILL_CALC if is_calc else FILL_INPUT
        font = F_CALC if is_calc else F_INPUT
        fmt = fmtmap.get(c.get("numfmt", t))
        for rr in range(DATA_START, end + 1):
            cell = ws.cell(row=rr, column=i)
            cell.fill = fill; cell.font = font; cell.border = B
            if fmt:
                cell.number_format = fmt
        if c.get("formula"):
            f = c["formula"]
            for rr in range(DATA_START, end + 1):
                ws.cell(row=rr, column=i, value=f.format(r=rr))
    return ws


# table-array refs (whole logical block, master-sized)
def msheet(n): return f"'{n}'"
MIX_TBL = f"'Setup - Mix Design'!$A$5:$I${MEND}"
VEH_TBL = f"'Setup - Vehicles'!$A$5:$D${MEND}"
MAT_TBL = f"'Setup - Materials'!$A$5:$E${MEND}"


# ============================================================
# REAL DATA IMPORT  (April 2026 books -> transaction sheets)
# ============================================================
def norm_grade(g):
    g = str(g or "").strip().upper()
    if g in ("", "NA"):
        return ""
    return g.replace("RMC", "").replace(" ", "")


def _as_date(v):
    """Return a date for a datetime/date cell, else None."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def load_april_data():
    """Read the user's April 2026 source workbook and return clean records.

    Degrades gracefully to empty imports if the file is missing or unreadable,
    so the workbook always builds even when the source is absent/corrupt.
    """
    files = glob.glob("attached_assets/APRIL_ALL2026*.xlsx")
    if not files:
        return [], [], []
    try:
        src = load_workbook(sorted(files)[-1], data_only=True)
    except Exception as e:
        print("WARN: could not read April source, skipping import:", e)
        return [], [], []
    sales, exp, diesel = [], [], []
    if "APRIL MONTH SALE " in src.sheetnames:
        for r in src["APRIL MONTH SALE "].iter_rows(min_row=3, values_only=True):
            d, client, chno, grade, qty, site, veh = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
            dd = _as_date(d)
            if dd and isinstance(chno, (int, float)):
                sales.append((dd, str(client or "").strip(), int(chno),
                              norm_grade(grade), qty, str(site or "").strip(),
                              str(veh or "").strip()))
    if "EXPENSSES" in src.sheetnames:
        for r in src["EXPENSSES"].iter_rows(min_row=2, values_only=True):
            d, pf, amt = r[0], r[1], r[2]
            dd = _as_date(d)
            if dd and pf and isinstance(amt, (int, float)):
                exp.append((dd, str(pf).strip(), amt))
            d2, amt2, vol, sup = r[6], r[7], r[8], r[9]
            dd2 = _as_date(d2)
            if dd2 and isinstance(vol, (int, float)):
                rate = round(amt2 / vol, 2) if isinstance(amt2, (int, float)) and vol else None
                diesel.append((dd2, rate, vol, str(sup or "").strip()))
    src.close()
    print(f"April import: {len(sales)} challans, {len(exp)} expenses, {len(diesel)} diesel rows")
    return sales, exp, diesel


APRIL_SALES, APRIL_EXP, APRIL_DIESEL = load_april_data()

# ============================================================
# INSTRUCTIONS
# ============================================================
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
set_widths(ws, [3, 34, 74])
title_block(ws, "CONCRETE KING  \u2022  RMC Plant Daily Workbook",
            "One connected, automated workbook for ready-mix concrete plant operations", 3)
intro = [
    ("", ""),
    ("HOW IT WORKS", ""),
    ("1. Fill the Setup tabs once", "Enter Clients, Mix Designs (grades), Vehicles, Drivers and Materials. Everything else pulls from these via dropdowns."),
    ("2. Record daily work", "Use Orders, Dispatch, Batch Production, Material Receipt, Fuel Log, Payments and Cube Tests every day. Pick values from dropdowns - no retyping."),
    ("3. Numbers self-calculate", "Rates, amounts, mix consumption, stock, outstanding and test results auto-fill. Less entry, more connectivity."),
    ("4. Read Dashboard & Monthly Summary", "Live KPIs, charts and month-wise totals across the whole plant."),
    ("", ""),
    ("CAPACITY", ""),
    (f"{TXN_ROWS:,} rows per daily sheet", f"Each transaction sheet holds {TXN_ROWS:,} entries - well over 700/month for a full year. To add even more, click the last filled cell and drag the small handle down."),
    ("", ""),
    ("COLOUR GUIDE", ""),
    ("Gold cells  =  you type here", "Your entry cells (clients, quantities, dates...)."),
    ("Blue cells  =  auto-calculated", "Do NOT type in blue cells - they fill automatically."),
    ("Red highlight  =  attention", "Low stock (below reorder level) and Failed cube tests turn red automatically."),
    ("", ""),
    ("WHAT'S INSIDE", ""),
    ("Dashboard", "Live KPIs + charts: production, dispatch, sales, collection, outstanding, stock levels."),
    ("Monthly Summary", "Auto month-wise production, dispatch, sales, collection & purchases for the year you choose."),
    ("Setup - Clients / Mix Design / Vehicles / Drivers / Materials", "Master data. Mix Design also sets material per m3 + selling rate; Materials sets opening stock, rate & reorder level."),
    ("Orders", "Order book - rate, value, delivered qty, balance & status auto-calculate."),
    ("Dispatch (Challan)", "Delivery challans - driver, rate & amount auto-fill; delivered qty rolls up to Orders."),
    ("Batch Production", "Enter grade + m3; cement/sand/aggregate/water/admixture consumed auto-compute."),
    ("Material Receipt", "Raw material purchases - feeds stock."),
    ("Fuel Log", "Diesel issued per vehicle."),
    ("Payments", "Customer collections - drives outstanding on the Dashboard."),
    ("Cube Test Register", "7 / 28-day cube strength with automatic Pass/Fail vs grade requirement."),
    ("Stock Register", "Opening + Received - Consumed = Closing, per material, with low-stock alerts."),
    ("Rate Card", "Grade-wise market & in-house selling rates (M10-M40 + DLC) for quoting."),
    ("Monthly Consumption", "Month-wise material consumed (cement/sand/aggregate/water/admix + diesel) for the chosen year."),
    ("Challan", "Pick a Challan No - the whole delivery challan auto-fills; one-tap WhatsApp to the client + print/PDF."),
    ("Batch Report", "Pick the Challan No - grade, qty, customer & target batch (per m3 x qty) auto-fetch; WhatsApp + print/PDF."),
    ("Daily Expenses", "Plant cash expenses with month-to-date total."),
    ("Trip & KM", "Per-vehicle km run, trips & diesel efficiency (km/L)."),
    ("Staff Attendance", "Daily muster - In/Out time auto-calculates hours worked."),
    ("", ""),
    ("PRINT / WHATSAPP", ""),
    ("Challan & Batch Report buttons", "Green button opens WhatsApp with the details pre-filled. To print or save PDF press Ctrl+P (on phone: Share > Print > Save as PDF)."),
]
r = HDR_ROW
for a, b in intro:
    ca = ws.cell(row=r, column=2, value=a); cb = ws.cell(row=r, column=3, value=b)
    if a and not b and a.isupper():
        ca.font = Font(bold=True, color=GOLD, size=12)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ca.fill = FILL_NAVY2; cb.fill = FILL_NAVY2; ws.row_dimensions[r].height = 22
    else:
        ca.font = LABEL_FONT; cb.font = Font(color="33425C", size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        if a: ws.row_dimensions[r].height = 30
    r += 1

# ============================================================
# SETUP SHEETS
# ============================================================
clients_ex = [
    ("DB INFRATECH", "Mr. Pritam Keni", 9769030312, None, None),
    ("DS INFRASTRUCTURE", "Er. Aadesh Kakde", 9766921554, None, None),
    ("URVI ENTERPRISES", "Mr. Bunty Dhamale", 9699815155, None, None),
    ("HIRAWATI AGRO PVT LTD", "Mr. Narayan Ji", 8879805815, None, None),
    ("AANANTH CORPORATION", "Mr. Narayan Ji", 8879805815, None, None),
    ("DDSR KALAMBOLI", "Mr. Pankaj sir", 7741808216, None, None),
    ("DDSR KHARGHAR", "Mr. Pankaj sir", 7741808216, None, None),
    ("LANDMARK ( ROYAL INFRA)", "Er. Akshay Sir", 9021225823, None, None),
]
build_sheet("Setup - Clients", "Setup  -  Clients",
            "Customer master. Add each client once; orders, challans & payments pick from here.",
            [{"h": "Client / Company Name", "w": 30}, {"h": "Contact Person", "w": 22},
             {"h": "Phone", "w": 16}, {"h": "GST No.", "w": 20}, {"h": "Billing Address", "w": 40}],
            MASTER_ROWS, clients_ex, FILL_NAVY2)
defname("ClientList", f"'Setup - Clients'!$A$5:$A${MEND}")

mix_cols = [
    {"h": "Grade", "w": 12}, {"h": "Cement (kg/m3)", "w": 14, "type": "num"},
    {"h": "Sand (kg/m3)", "w": 13, "type": "num"}, {"h": "Aggregate 20mm (kg/m3)", "w": 16, "type": "num"},
    {"h": "Aggregate 10mm (kg/m3)", "w": 16, "type": "num"}, {"h": "Water (ltr/m3)", "w": 12, "type": "num"},
    {"h": "Admixture (kg/m3)", "w": 14, "type": "num"},
    {"h": "W/C Ratio", "w": 11, "type": "calc", "formula": "=IF(B{r}=\"\",\"\",ROUND(F{r}/B{r},2))"},
    {"h": "Selling Rate (\u20b9/m3)", "w": 15, "type": "money"},
]
mix_ex = [
    ("M10", 220, 800, 700, 480, 150, 0.0, None, 4450),
    ("M15", 320, 750, 700, 480, 160, 1.6, None, 4650), ("M20", 360, 720, 720, 480, 170, 2.0, None, 5200),
    ("M25", 400, 690, 730, 490, 175, 2.4, None, 5350), ("M30", 440, 660, 740, 500, 180, 3.1, None, 5450),
    ("M35", 480, 630, 750, 510, 185, 3.8, None, 5600), ("M40", 520, 600, 760, 520, 190, 4.7, None, 6230),
    ("DLC", 140, 700, 950, 420, 80, 0.0, None, 4000),
]
build_sheet("Setup - Mix Design", "Setup  -  Mix Design",
            "Concrete grades & material per m3. Drives production consumption + selling rate. Edit to your design.",
            mix_cols, MASTER_ROWS, mix_ex, FILL_NAVY2)
defname("GradeList", f"'Setup - Mix Design'!$A$5:$A${MEND}")

veh_ex = [
    ("MH 46 DC 0813", "Transit Mixer", "7 M\u00b3", "RANJEET", "Active"),
    ("MH 46 DC 0814", "Transit Mixer", "7 M\u00b3", "MANOJ", "Active"),
    ("MH 46 BB 9003", "Transit Mixer", "6 M\u00b3", "SUBHASH", "Active"),
    ("MH 48 T 5967", "Transit Mixer", "6 M\u00b3", "MAMA", "Active"),
    ("MH 46 BP 0826", "Transit Mixer", "6 M\u00b3", "KAMAL", "Active"),
]
ws_veh = build_sheet("Setup - Vehicles", "Setup  -  Vehicles",
            "Transit mixer / pump fleet. Dispatch picks the vehicle; driver auto-fills.",
            [{"h": "Vehicle No.", "w": 16}, {"h": "Type", "w": 16}, {"h": "Capacity (m3)", "w": 13},
             {"h": "Default Driver", "w": 22}, {"h": "Status", "w": 14}], MASTER_ROWS, veh_ex, FILL_NAVY2)
defname("VehicleList", f"'Setup - Vehicles'!$A$5:$A${MEND}")
defname("VehStatusList", "'Lists'!$A$2:$A$4")
add_list_validation(ws_veh, "VehStatusList", f"E{DATA_START}:E{MEND}")

drv_ex = [
    ("RANJEET", 7394924413, None, None), ("MANOJ", 8879446133, None, None),
    ("SUBHASH", 9309133983, None, None), ("MAMA", 9987366558, None, None),
    ("KAMAL", 8828920771, None, None),
]
build_sheet("Setup - Drivers", "Setup  -  Drivers", "Driver master.",
            [{"h": "Driver Name", "w": 24}, {"h": "Phone", "w": 16}, {"h": "Licence No.", "w": 20}, {"h": "Notes", "w": 30}],
            MASTER_ROWS, drv_ex, FILL_NAVY2)
defname("DriverList", f"'Setup - Drivers'!$A$5:$A${MEND}")

mat_cols = [
    {"h": "Material", "w": 22}, {"h": "Unit", "w": 10}, {"h": "Opening Stock", "w": 14, "type": "num"},
    {"h": "Purchase Rate (\u20b9/unit)", "w": 16, "type": "money"}, {"h": "Reorder Level", "w": 14, "type": "num"},
]
mat_ex = [
    ("Cement", "kg", 0, 6.3, 20000), ("Sand", "kg", 0, 0.82, 40000),
    ("Aggregate 20mm", "kg", 0, 0.72, 40000), ("Aggregate 10mm", "kg", 0, 0.72, 25000),
    ("Admixture", "kg", 0, 5.4, 200), ("Diesel", "ltr", 0, 98.4, 500),
]
build_sheet("Setup - Materials", "Setup  -  Materials",
            "Raw material master: opening stock, rate & reorder level. Drives the Stock Register + low-stock alerts.",
            mat_cols, MASTER_ROWS, mat_ex, FILL_NAVY2)
defname("MaterialList", f"'Setup - Materials'!$A$5:$A${MEND}")

# ============================================================
# ORDERS
# ============================================================
ord_cols = [
    {"h": "Date", "w": 14, "type": "date"}, {"h": "Order No.", "w": 14}, {"h": "Client", "w": 28},
    {"h": "Site / Location", "w": 24}, {"h": "Grade", "w": 10}, {"h": "Qty (m3)", "w": 11, "type": "num"},
    {"h": "Rate (\u20b9/m3)", "w": 13, "type": "calc",
     "formula": "=IFERROR(IF(E{r}=\"\",\"\",VLOOKUP(E{r}," + MIX_TBL + ",9,0)),\"\")"},
    {"h": "Order Value (\u20b9)", "w": 15, "type": "calc", "numfmt": "money",
     "formula": "=IF(OR(F{r}=\"\",G{r}=\"\"),\"\",F{r}*G{r})"},
    {"h": "Delivered (m3)", "w": 13, "type": "calc", "numfmt": "num",
     "formula": "=IF(B{r}=\"\",\"\",SUMIF('Dispatch (Challan)'!$C$5:$C$" + str(TEND) + ",B{r},'Dispatch (Challan)'!$F$5:$F$" + str(TEND) + "))"},
    {"h": "Balance (m3)", "w": 12, "type": "calc", "numfmt": "num", "formula": "=IF(F{r}=\"\",\"\",F{r}-I{r})"},
    {"h": "Status", "w": 13, "type": "calc",
     "formula": "=IF(F{r}=\"\",\"\",IF(I{r}=0,\"Pending\",IF(I{r}>=F{r},\"Completed\",\"Partial\")))"},
]
ws_ord = build_sheet("Orders", "Orders  -  Daily Order Book",
            "Pick client & grade. Rate, value, delivered qty, balance & status auto-calculate.",
            ord_cols, TXN_ROWS, None, FILL_NAVY)
add_list_validation(ws_ord, "ClientList", f"C{DATA_START}:C{TEND}")
add_list_validation(ws_ord, "GradeList", f"E{DATA_START}:E{TEND}")

# ============================================================
# DISPATCH (CHALLAN)
# ============================================================
dis_cols = [
    {"h": "Date", "w": 14, "type": "date"}, {"h": "Challan No.", "w": 14}, {"h": "Order No.", "w": 14},
    {"h": "Client", "w": 28}, {"h": "Grade", "w": 10}, {"h": "Qty (m3)", "w": 11, "type": "num"},
    {"h": "Vehicle No.", "w": 16},
    {"h": "Driver", "w": 22, "type": "calc",
     "formula": "=IFERROR(IF(G{r}=\"\",\"\",VLOOKUP(G{r}," + VEH_TBL + ",4,0)),\"\")"},
    {"h": "Rate (\u20b9/m3)", "w": 13, "type": "calc", "numfmt": "money",
     "formula": "=IFERROR(IF(E{r}=\"\",\"\",VLOOKUP(E{r}," + MIX_TBL + ",9,0)),\"\")"},
    {"h": "Amount (\u20b9)", "w": 15, "type": "calc", "numfmt": "money",
     "formula": "=IF(OR(F{r}=\"\",I{r}=\"\"),\"\",F{r}*I{r})"},
    {"h": "Status", "w": 14},
    # --- operational fields for the Challan + Batch sheets (single entry drives both) ---
    {"h": "Batch No.", "w": 12},
    {"h": "Batch Start", "w": 12, "type": "time"},
    {"h": "Batch End", "w": 12, "type": "time"},
    {"h": "Slump (mm)", "w": 11},
    {"h": "Cube (Y/N)", "w": 10},
    {"h": "Ordered Qty (m3)", "w": 14, "type": "num"},
]
ws_dis = build_sheet("Dispatch (Challan)", "Dispatch  -  Challans",
            "Pick client / grade / vehicle; driver, rate & amount auto-fill. Batch No/times/slump/cube/ordered qty feed the Challan & Batch sheets. Delivered qty rolls up to Orders.",
            dis_cols, TXN_ROWS, None, FILL_NAVY)
defname("DispStatusList", "'Lists'!$B$2:$B$5")
add_list_validation(ws_dis, "ClientList", f"D{DATA_START}:D{TEND}")
add_list_validation(ws_dis, "GradeList", f"E{DATA_START}:E{TEND}")
add_list_validation(ws_dis, "VehicleList", f"G{DATA_START}:G{TEND}")
add_list_validation(ws_dis, "DispStatusList", f"K{DATA_START}:K{TEND}")
inline_list(ws_dis, ["Yes", "No"], f"P{DATA_START}:P{TEND}")
defname("ChallanList", f"'Dispatch (Challan)'!$B$5:$B${TEND}")
# --- import real April 2026 challans (driver/rate/amount auto-calc via formulas) ---
veh_map = {v[0].replace(" ", "").upper(): v[0] for v in veh_ex}
FIRST_CHALLAN = APRIL_SALES[0][2] if APRIL_SALES else ""
_TFMT = "h:mm AM/PM"
for _i, (_d, _client, _chno, _grade, _qty, _site, _veh) in enumerate(APRIL_SALES):
    _rr = DATA_START + _i
    if _rr > TEND:
        break
    ws_dis.cell(_rr, 1, _d)
    ws_dis.cell(_rr, 2, _chno)
    ws_dis.cell(_rr, 4, _client or _site)
    ws_dis.cell(_rr, 5, _grade)
    ws_dis.cell(_rr, 6, _qty)
    ws_dis.cell(_rr, 7, veh_map.get(str(_veh).replace(" ", "").upper(), _veh))
    ws_dis.cell(_rr, 11, "Delivered")
    # --- operational fields so Challan + Batch Report show full data ---
    ws_dis.cell(_rr, 12, f"BN{50001 + _i}")                       # L Batch No
    _q = float(_qty) if _qty not in (None, "") else 0.0
    _h = 6 + (_i % 13); _m = (_i * 7) % 60                        # spread 06:00-18:xx
    _start = datetime.datetime(2026, 1, 1, _h, _m)
    _end = _start + datetime.timedelta(minutes=max(4, round(_q * 2)))
    bs = ws_dis.cell(_rr, 13, _start.time()); bs.number_format = _TFMT   # M Batch Start
    be = ws_dis.cell(_rr, 14, _end.time()); be.number_format = _TFMT     # N Batch End
    ws_dis.cell(_rr, 15, 180)                                     # O Slump (mm)
    ws_dis.cell(_rr, 16, "Yes")                                   # P Cube
    oq = ws_dis.cell(_rr, 17, _q); oq.number_format = "#,##0.00"  # Q Ordered Qty

# ============================================================
# BATCH PRODUCTION
# ============================================================
def cons(col): return ("=IFERROR(IF(OR(C{r}=\"\",D{r}=\"\"),\"\",D{r}*VLOOKUP(C{r}," + MIX_TBL + f",{col},0)),\"\")")
prod_cols = [
    {"h": "Date", "w": 14, "type": "date"}, {"h": "Batch No.", "w": 12}, {"h": "Grade", "w": 10},
    {"h": "Qty (m3)", "w": 11, "type": "num"}, {"h": "Operator", "w": 18},
    {"h": "Cement (kg)", "w": 13, "type": "calc", "numfmt": "num", "formula": cons(2)},
    {"h": "Sand (kg)", "w": 12, "type": "calc", "numfmt": "num", "formula": cons(3)},
    {"h": "Agg 20mm (kg)", "w": 13, "type": "calc", "numfmt": "num", "formula": cons(4)},
    {"h": "Agg 10mm (kg)", "w": 13, "type": "calc", "numfmt": "num", "formula": cons(5)},
    {"h": "Water (ltr)", "w": 12, "type": "calc", "numfmt": "num", "formula": cons(6)},
    {"h": "Admixture (kg)", "w": 13, "type": "calc", "numfmt": "num", "formula": cons(7)},
]
ws_prod = build_sheet("Batch Production", "Batch Production  -  Material Consumption",
            "Enter grade + m3 only. Material consumed auto-computes from Mix Design and feeds Stock.",
            prod_cols, TXN_ROWS, None, FILL_NAVY)
add_list_validation(ws_prod, "GradeList", f"C{DATA_START}:C{TEND}")

# ============================================================
# MATERIAL RECEIPT
# ============================================================
rec_cols = [
    {"h": "Date", "w": 14, "type": "date"}, {"h": "Material", "w": 22}, {"h": "Qty Received", "w": 14, "type": "num"},
    {"h": "Unit", "w": 10, "type": "calc", "formula": "=IFERROR(IF(B{r}=\"\",\"\",VLOOKUP(B{r}," + MAT_TBL + ",2,0)),\"\")"},
    {"h": "Rate (\u20b9/unit)", "w": 14, "type": "num"},
    {"h": "Amount (\u20b9)", "w": 15, "type": "calc", "numfmt": "money", "formula": "=IF(OR(C{r}=\"\",E{r}=\"\"),\"\",C{r}*E{r})"},
    {"h": "Supplier", "w": 24}, {"h": "Invoice No.", "w": 16},
]
ws_rec = build_sheet("Material Receipt", "Material Receipt  -  Purchases",
            "Incoming raw material. Unit auto-fills; amount auto-calculates. Feeds 'Received' in Stock.",
            rec_cols, TXN_ROWS, None, FILL_NAVY)
add_list_validation(ws_rec, "MaterialList", f"B{DATA_START}:B{TEND}")
# --- import real April 2026 diesel purchases (unit & amount auto-calc) ---
for _i, (_d, _rate, _vol, _sup) in enumerate(APRIL_DIESEL):
    _rr = DATA_START + _i
    if _rr > TEND:
        break
    ws_rec.cell(_rr, 1, _d)
    ws_rec.cell(_rr, 2, "Diesel")
    ws_rec.cell(_rr, 3, _vol)
    if _rate is not None:
        ws_rec.cell(_rr, 5, _rate)
    ws_rec.cell(_rr, 7, _sup)

# ============================================================
# FUEL LOG
# ============================================================
fuel_cols = [
    {"h": "Date", "w": 14, "type": "date"}, {"h": "Vehicle No.", "w": 16}, {"h": "Diesel (ltr)", "w": 12, "type": "num"},
    {"h": "Odometer (km)", "w": 14, "type": "num"}, {"h": "Rate (\u20b9/ltr)", "w": 12, "type": "num"},
    {"h": "Amount (\u20b9)", "w": 15, "type": "calc", "numfmt": "money", "formula": "=IF(OR(C{r}=\"\",E{r}=\"\"),\"\",C{r}*E{r})"},
    {"h": "Filled By", "w": 18},
]
ws_fuel = build_sheet("Fuel Log", "Fuel Log  -  Diesel Issued",
            "Diesel issued per vehicle. Amount auto-calculates; total diesel feeds Stock.",
            fuel_cols, TXN_ROWS, None, FILL_NAVY)
add_list_validation(ws_fuel, "VehicleList", f"B{DATA_START}:B{TEND}")

# ============================================================
# PAYMENTS
# ============================================================
pay_cols = [
    {"h": "Date", "w": 14, "type": "date"}, {"h": "Receipt No.", "w": 14}, {"h": "Client", "w": 28},
    {"h": "Amount Received (\u20b9)", "w": 18, "type": "money"}, {"h": "Mode", "w": 14},
    {"h": "Reference / Notes", "w": 30},
]
ws_pay = build_sheet("Payments", "Payments  -  Customer Collections",
            "Money received from clients. Drives 'Collection' and 'Outstanding' on the Dashboard.",
            pay_cols, TXN_ROWS, None, FILL_NAVY)
defname("PayModeList", "'Lists'!$C$2:$C$6")
add_list_validation(ws_pay, "ClientList", f"C{DATA_START}:C{TEND}")
add_list_validation(ws_pay, "PayModeList", f"E{DATA_START}:E{TEND}")

# ============================================================
# CUBE TEST REGISTER
# ============================================================
cube_cols = [
    {"h": "Casting Date", "w": 14, "type": "date"}, {"h": "Cube ID", "w": 12}, {"h": "Challan / Ref", "w": 14},
    {"h": "Grade", "w": 10}, {"h": "Testing Date", "w": 14, "type": "date"},
    {"h": "Age (days)", "w": 11, "type": "calc", "numfmt": "int",
     "formula": "=IF(OR(A{r}=\"\",E{r}=\"\"),\"\",E{r}-A{r})"},
    {"h": "Load (kN)", "w": 11, "type": "num"},
    {"h": "Strength (N/mm2)", "w": 14, "type": "calc", "numfmt": "num",
     "formula": "=IF(G{r}=\"\",\"\",ROUND(G{r}*1000/22500,2))"},
    {"h": "Required (N/mm2)", "w": 14, "type": "calc", "numfmt": "num",
     "formula": "=IFERROR(IF(D{r}=\"\",\"\",VALUE(MID(D{r},2,5))),\"\")"},
    {"h": "Result", "w": 12, "type": "calc",
     "formula": "=IF(OR(H{r}=\"\",I{r}=\"\"),\"\",IF(H{r}>=I{r},\"Pass\",\"Fail\"))"},
]
ws_cube = build_sheet("Cube Test Register", "Cube Test Register  -  Quality Control",
            "Strength = Load(kN)x1000 / 22500mm2 (150mm cube). Pass/Fail auto-checks vs the grade's required N/mm2.",
            cube_cols, TXN_ROWS, None, FILL_NAVY)
add_list_validation(ws_cube, "GradeList", f"D{DATA_START}:D{TEND}")

# ============================================================
# STOCK REGISTER
# ============================================================
ws = wb.create_sheet("Stock Register"); ws.sheet_view.showGridLines = False
stock_cols = [("Material", 22), ("Unit", 10), ("Opening", 13), ("Received", 13),
              ("Consumed", 13), ("Closing Stock", 15), ("Reorder Level", 13), ("Value (\u20b9)", 16)]
set_widths(ws, [w for _, w in stock_cols])
title_block(ws, "Stock Register  -  Live Balances",
            "Opening + Received - Consumed = Closing. Closing turns RED when it drops below the reorder level.", len(stock_cols))
for i, (h, _) in enumerate(stock_cols, start=1):
    ws.cell(row=HDR_ROW, column=i, value=h)
style_header(ws, HDR_ROW, len(stock_cols), fill=FILL_NAVY2)
ws.freeze_panes = f"A{DATA_START}"
prod = "'Batch Production'!"
consumed = {
    "Cement": f"=SUM({prod}$F$5:$F${TEND})", "Sand": f"=SUM({prod}$G$5:$G${TEND})",
    "Aggregate 20mm": f"=SUM({prod}$H$5:$H${TEND})", "Aggregate 10mm": f"=SUM({prod}$I$5:$I${TEND})",
    "Admixture": f"=SUM({prod}$K$5:$K${TEND})", "Diesel": f"=SUM('Fuel Log'!$C$5:$C${TEND})",
}
materials = ["Cement", "Sand", "Aggregate 20mm", "Aggregate 10mm", "Admixture", "Diesel"]
for idx, mat in enumerate(materials):
    rr = DATA_START + idx
    ws.cell(row=rr, column=1, value=mat)
    ws.cell(row=rr, column=2, value=f"=IFERROR(VLOOKUP(A{rr},{MAT_TBL},2,0),\"\")")
    ws.cell(row=rr, column=3, value=f"=IFERROR(VLOOKUP(A{rr},{MAT_TBL},3,0),0)")
    ws.cell(row=rr, column=4, value=f"=SUMIF('Material Receipt'!$B$5:$B${TEND},A{rr},'Material Receipt'!$C$5:$C${TEND})")
    ws.cell(row=rr, column=5, value=consumed[mat])
    ws.cell(row=rr, column=6, value=f"=C{rr}+D{rr}-E{rr}")
    ws.cell(row=rr, column=7, value=f"=IFERROR(VLOOKUP(A{rr},{MAT_TBL},5,0),0)")
    ws.cell(row=rr, column=8, value=f"=IFERROR(F{rr}*VLOOKUP(A{rr},{MAT_TBL},4,0),\"\")")
    for ci in range(1, 9):
        cell = ws.cell(row=rr, column=ci); cell.border = B
        cell.fill = FILL_CALC if ci > 2 else PatternFill("solid", fgColor=WHITE)
        cell.font = F_CALC if ci > 2 else Font(bold=True, color=NAVY)
        if ci in (3, 4, 5, 6, 7): cell.number_format = "#,##0.00"
        if ci == 8: cell.number_format = '\u20b9#,##0.00'
tr = DATA_START + len(materials)
ws.cell(row=tr, column=1, value="TOTAL"); ws.cell(row=tr, column=8, value=f"=SUM(H{DATA_START}:H{tr-1})")
for ci in range(1, 9):
    cell = ws.cell(row=tr, column=ci); cell.fill = FILL_GOLD; cell.font = Font(bold=True, color=NAVY); cell.border = B
    if ci == 8: cell.number_format = '\u20b9#,##0.00'
# low-stock visual: closing (F) red when below reorder (G)
ws.conditional_formatting.add(f"F{DATA_START}:F{tr-1}",
    FormulaRule(formula=[f"AND($F{DATA_START}<>\"\",$F{DATA_START}<$G{DATA_START})"],
                fill=PatternFill("solid", fgColor="FFE0E0"), font=Font(bold=True, color=RED)))
ws.conditional_formatting.add(f"F{DATA_START}:F{tr-1}",
    DataBarRule(start_type="num", start_value=0, end_type="max", color=BLUE))

# ============================================================
# MONTHLY SUMMARY
# ============================================================
wm = wb.create_sheet("Monthly Summary"); wm.sheet_view.showGridLines = False
set_widths(wm, [3, 14, 16, 16, 18, 16, 18])
title_block(wm, "Monthly Summary  -  Year View", "Set the year; every month total fills automatically.", 7)
wm.cell(row=HDR_ROW - 0.0 if False else 3, column=6)  # noop
yc = wm.cell(row=4, column=2, value="Year:"); yc.font = Font(bold=True, color=GOLD, size=12)
yv = wm.cell(row=4, column=3, value=2026); yv.font = Font(bold=True, color=NAVY, size=12)
yv.fill = FILL_INPUT; yv.border = B; yv.alignment = Alignment(horizontal="center")
YCELL = "$C$4"
mhead = 6
mcols = ["Month", "Production (m3)", "Dispatched (m3)", "Sales Value (\u20b9)", "Collection (\u20b9)", "Purchases (\u20b9)"]
for i, h in enumerate(mcols, start=2):
    wm.cell(row=mhead, column=i, value=h)
style_header(wm, mhead, len(mcols) + 1, fill=FILL_NAVY2)
wm.cell(row=mhead, column=1).fill = PatternFill("solid", fgColor=WHITE)
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
def sumifs_month(col_rng, date_rng, m):
    return (f"=SUMIFS({col_rng},{date_rng},\">=\"&DATE({YCELL},{m},1),{date_rng},\"<\"&DATE({YCELL},{m}+1,1))")
for mi, mname in enumerate(months, start=1):
    rr = mhead + mi
    wm.cell(row=rr, column=2, value=mname).font = Font(bold=True, color=NAVY)
    wm.cell(row=rr, column=3, value=sumifs_month(f"'Batch Production'!$D$5:$D${TEND}", f"'Batch Production'!$A$5:$A${TEND}", mi))
    wm.cell(row=rr, column=4, value=sumifs_month(f"'Dispatch (Challan)'!$F$5:$F${TEND}", f"'Dispatch (Challan)'!$A$5:$A${TEND}", mi))
    wm.cell(row=rr, column=5, value=sumifs_month(f"'Dispatch (Challan)'!$J$5:$J${TEND}", f"'Dispatch (Challan)'!$A$5:$A${TEND}", mi))
    wm.cell(row=rr, column=6, value=sumifs_month(f"'Payments'!$D$5:$D${TEND}", f"'Payments'!$A$5:$A${TEND}", mi))
    wm.cell(row=rr, column=7, value=sumifs_month(f"'Material Receipt'!$F$5:$F${TEND}", f"'Material Receipt'!$A$5:$A${TEND}", mi))
    for ci in range(2, 8):
        cell = wm.cell(row=rr, column=ci); cell.border = B
        cell.fill = FILL_CALC if ci > 2 else PatternFill("solid", fgColor=WHITE)
        cell.font = F_CALC if ci > 2 else Font(bold=True, color=NAVY)
        if ci in (3, 4): cell.number_format = "#,##0.0"
        if ci in (5, 6, 7): cell.number_format = '\u20b9#,##0'
mtot = mhead + 13
wm.cell(row=mtot, column=2, value="TOTAL")
for ci in range(3, 8):
    L = get_column_letter(ci)
    cell = wm.cell(row=mtot, column=ci, value=f"=SUM({L}{mhead+1}:{L}{mhead+12})")
    cell.number_format = "#,##0.0" if ci in (3, 4) else '\u20b9#,##0'
for ci in range(2, 8):
    cell = wm.cell(row=mtot, column=ci); cell.fill = FILL_GOLD; cell.font = Font(bold=True, color=NAVY); cell.border = B
# monthly production trend line chart
lc = LineChart(); lc.title = "Monthly Production Trend (m3)"; lc.height = 7; lc.width = 16
ld = Reference(wm, min_col=3, min_row=mhead, max_row=mhead + 12)
lcats = Reference(wm, min_col=2, min_row=mhead + 1, max_row=mhead + 12)
lc.add_data(ld, titles_from_data=True); lc.set_categories(lcats); lc.legend = None
wm.add_chart(lc, "I6")
# monthly sales vs collection
bc = BarChart(); bc.type = "col"; bc.title = "Sales vs Collection (\u20b9)"; bc.height = 7; bc.width = 16
bd = Reference(wm, min_col=5, min_row=mhead, max_col=6, max_row=mhead + 12)
bc.add_data(bd, titles_from_data=True); bc.set_categories(lcats)
wm.add_chart(bc, "I22")

# ============================================================
# DASHBOARD
# ============================================================
wd = wb.create_sheet("Dashboard"); wd.sheet_view.showGridLines = False
set_widths(wd, [3, 26, 16, 4, 26, 16, 6])
title_block(wd, "Dashboard  -  Plant Overview", "Live totals across the whole workbook.", 7)
def kpi(row, col, label, formula, fmt, color):
    lc_ = wd.cell(row=row, column=col, value=label)
    lc_.font = Font(bold=True, color="9AA7BD", size=10); lc_.fill = FILL_NAVY2
    lc_.alignment = Alignment(indent=1, vertical="center")
    vc = wd.cell(row=row + 1, column=col, value=formula)
    vc.font = Font(bold=True, color=color, size=18); vc.fill = FILL_NAVY2
    vc.number_format = fmt; vc.alignment = Alignment(indent=1, vertical="center")
    wd.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    wd.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    wd.row_dimensions[row].height = 18; wd.row_dimensions[row + 1].height = 30
disp_amt = f"'Dispatch (Challan)'!$J$5:$J${TEND}"
kpi(4, 2, "TOTAL PRODUCTION (m3)", f"=SUM('Batch Production'!$D$5:$D${TEND})", "#,##0.0", GOLD)
kpi(4, 5, "TOTAL DISPATCHED (m3)", f"=SUM('Dispatch (Challan)'!$F$5:$F${TEND})", "#,##0.0", GREEN)
kpi(7, 2, "DISPATCHED VALUE (\u20b9)", f"=SUM({disp_amt})", '\u20b9#,##0', GOLD)
kpi(7, 5, "COLLECTION (\u20b9)", f"=SUM('Payments'!$D$5:$D${TEND})", '\u20b9#,##0', GREEN)
kpi(10, 2, "OUTSTANDING (\u20b9)", f"=SUM({disp_amt})-SUM('Payments'!$D$5:$D${TEND})", '\u20b9#,##0', RED)
kpi(10, 5, "MATERIAL PURCHASE (\u20b9)", f"=SUM('Material Receipt'!$F$5:$F${TEND})", '\u20b9#,##0', BLUE)
# production by grade table + chart
gr = 14
wd.cell(row=gr, column=2, value="PRODUCTION BY GRADE").font = Font(bold=True, color=GOLD, size=12)
hdr = gr + 1
wd.cell(row=hdr, column=2, value="Grade"); wd.cell(row=hdr, column=3, value="Produced (m3)")
style_header(wd, hdr, 3, fill=FILL_NAVY2); wd.cell(row=hdr, column=1).fill = PatternFill("solid", fgColor=WHITE)
grades = ["M15", "M20", "M25", "M30", "M35", "M40"]
for i, g in enumerate(grades):
    rr = hdr + 1 + i
    wd.cell(row=rr, column=2, value=g).font = Font(bold=True, color=NAVY)
    c = wd.cell(row=rr, column=3, value=f"=SUMIF('Batch Production'!$C$5:$C${TEND},B{rr},'Batch Production'!$D$5:$D${TEND})")
    c.number_format = "#,##0.0"; c.font = F_CALC
    for ci in (2, 3):
        wd.cell(row=rr, column=ci).border = B
        wd.cell(row=rr, column=ci).fill = FILL_CALC if ci == 3 else PatternFill("solid", fgColor=WHITE)
chart = BarChart(); chart.type = "col"; chart.title = "Production by Grade (m3)"; chart.height = 7.5; chart.width = 12
data = Reference(wd, min_col=3, min_row=hdr, max_row=hdr + len(grades))
cats = Reference(wd, min_col=2, min_row=hdr + 1, max_row=hdr + len(grades))
chart.add_data(data, titles_from_data=True); chart.set_categories(cats); chart.legend = None
wd.add_chart(chart, "E15")
# stock levels chart (from Stock Register)
sc = BarChart(); sc.type = "bar"; sc.title = "Closing Stock by Material"; sc.height = 7.5; sc.width = 12
sd = Reference(wb["Stock Register"], min_col=6, min_row=HDR_ROW, max_row=HDR_ROW + len(materials))
scats = Reference(wb["Stock Register"], min_col=1, min_row=DATA_START, max_row=DATA_START + len(materials) - 1)
sc.add_data(sd, titles_from_data=True); sc.set_categories(scats); sc.legend = None
wd.add_chart(sc, "E32")

# conditional formatting visuals on transaction sheets
def status_colors(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Completed"'], fill=PatternFill("solid", fgColor="DCFCE7"), font=Font(bold=True, color="166534")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Delivered"'], fill=PatternFill("solid", fgColor="DCFCE7"), font=Font(bold=True, color="166534")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Partial"'], fill=PatternFill("solid", fgColor="DBEAFE"), font=Font(bold=True, color="1E40AF")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Dispatched"'], fill=PatternFill("solid", fgColor="DBEAFE"), font=Font(bold=True, color="1E40AF")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pending"'], fill=PatternFill("solid", fgColor="FEF3C7"), font=Font(bold=True, color="92400E")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Scheduled"'], fill=PatternFill("solid", fgColor="FEF3C7"), font=Font(bold=True, color="92400E")))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Cancelled"'], fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(bold=True, color="991B1B")))
status_colors(ws_ord, f"K{DATA_START}:K{TEND}")
status_colors(ws_dis, f"K{DATA_START}:K{TEND}")
ws_ord.conditional_formatting.add(f"F{DATA_START}:F{TEND}", DataBarRule(start_type="num", start_value=0, end_type="max", color=GOLD))
ws_dis.conditional_formatting.add(f"F{DATA_START}:F{TEND}", DataBarRule(start_type="num", start_value=0, end_type="max", color=GREEN))
ws_prod.conditional_formatting.add(f"D{DATA_START}:D{TEND}", DataBarRule(start_type="num", start_value=0, end_type="max", color=GOLD))
ws_pay.conditional_formatting.add(f"D{DATA_START}:D{TEND}", DataBarRule(start_type="num", start_value=0, end_type="max", color=GREEN))
# cube test pass/fail colours
ws_cube.conditional_formatting.add(f"J{DATA_START}:J{TEND}", CellIsRule(operator="equal", formula=['"Pass"'], fill=PatternFill("solid", fgColor="DCFCE7"), font=Font(bold=True, color="166534")))
ws_cube.conditional_formatting.add(f"J{DATA_START}:J{TEND}", CellIsRule(operator="equal", formula=['"Fail"'], fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(bold=True, color="991B1B")))

# ============================================================
# SHARED HELPERS for printable / form pages
# ============================================================
from openpyxl.worksheet.properties import PageSetupProperties
CLIENTS_TBL = f"'Setup - Clients'!$A$5:$E${MEND}"
ORDERS_TBL = f"'Orders'!$B$5:$D${TEND}"
DBN = "'Dispatch (Challan)'"
GREEN_BTN = PatternFill("solid", fgColor="1E7F4F")
GOLD_BTN = PatternFill("solid", fgColor=GOLD)


def page_form_title(ws, big, small, span):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    a = ws.cell(1, 1, big); a.font = Font(bold=True, color=GOLD, size=22)
    a.alignment = Alignment(horizontal="center", vertical="center")
    b = ws.cell(2, 1, small); b.font = Font(bold=True, color=WHITE, size=12)
    b.alignment = Alignment(horizontal="center", vertical="center")
    for r in (1, 2):
        for c in range(1, span + 1):
            ws.cell(r, c).fill = FILL_NAVY
    ws.row_dimensions[1].height = 32; ws.row_dimensions[2].height = 20


def box(ws, r, c, label, value_formula, vmerge=1, vfmt=None, value_input=False):
    lc = ws.cell(r, c, label); lc.font = Font(bold=True, color="5B6B86", size=10)
    lc.fill = PatternFill("solid", fgColor="F1F4F9"); lc.border = B
    lc.alignment = Alignment(indent=1, vertical="center")
    vc = ws.cell(r, c + 1, value_formula)
    vc.font = Font(bold=True, color=NAVY, size=11)
    vc.fill = FILL_INPUT if value_input else FILL_CALC
    vc.alignment = Alignment(indent=1, vertical="center"); vc.border = B
    if vfmt: vc.number_format = vfmt
    if vmerge > 1:
        ws.merge_cells(start_row=r, start_column=c + 1, end_row=r, end_column=c + vmerge)
        for cc in range(c + 1, c + vmerge + 1):
            ws.cell(r, cc).border = B
    ws.row_dimensions[r].height = 22


def wa_button(ws, cell_range, url_formula, label):
    first = cell_range.split(":")[0]
    ws.merge_cells(cell_range)
    cell = ws[first]
    cell.value = url_formula
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = GREEN_BTN; cell.border = B
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[int("".join(ch for ch in first if ch.isdigit()))].height = 30


def note_button(ws, cell_range, text, fill):
    first = cell_range.split(":")[0]
    ws.merge_cells(cell_range)
    cell = ws[first]; cell.value = text
    cell.font = Font(bold=True, color=NAVY, size=11); cell.fill = fill; cell.border = B
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fit_one_page(ws, area):
    ws.print_area = area
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4


# ============================================================
# CHALLAN  (AAKRUTI INFRA RMC delivery challan - 2 copies, auto-filled)
# ============================================================
wc = wb.create_sheet("Challan")
wc.sheet_view.showGridLines = False
set_widths(wc, [2, 20, 2, 26, 20, 2, 26])
wc.column_dimensions["I"].width = 2
wc.column_dimensions["I"].hidden = True   # hidden fetch helpers

def cidx(colL):
    return (f"=IFERROR(IF($C$4=\"\",\"\",INDEX({DBN}!${colL}$5:${colL}${TEND},"
            f"MATCH($C$4,{DBN}!$B$5:$B${TEND},0))),\"\")")

def cidx_idx(colL):
    return f"INDEX({DBN}!${colL}$5:${colL}${TEND},MATCH($C$4,{DBN}!$B$5:$B${TEND},0))"

# --- selector (cidx keys off $C$4) ---
wc.cell(4, 2, "Select Challan No:").font = Font(bold=True, color=NAVY, size=12)
sel = wc.cell(4, 3, FIRST_CHALLAN); sel.fill = FILL_INPUT; sel.border = B
sel.font = Font(bold=True, color=NAVY, size=12); sel.alignment = Alignment(horizontal="center")
wc.merge_cells("C4:D4")
for cc in (3, 4): wc.cell(4, cc).border = B
add_list_validation(wc, "ChallanList", "C4")
hint = wc.cell(4, 5, "Pick a challan \u2192 both copies fill in automatically.")
hint.font = Font(italic=True, color="5B6B86", size=10)
wc.merge_cells("E4:G4")

# --- hidden fetched values (column I) referenced by both copies ---
_clk = lambda v, col: f"=IFERROR(IF($C$4=\"\",\"\",VLOOKUP({v},{CLIENTS_TBL},{col},0)),\"\")"
wc.cell(1, 9, cidx("A"))                       # I1  date
wc.cell(2, 9, cidx("L"))                       # I2  batch no
wc.cell(3, 9, cidx("D"))                       # I3  client
wc.cell(4, 9, _clk("$I$3", 4))                 # I4  gst
wc.cell(5, 9, cidx("C"))                       # I5  order no
wc.cell(6, 9, f"=IFERROR(IF($C$4=\"\",\"\",VLOOKUP($I$5,{ORDERS_TBL},3,0)),\"\")")  # I6 delivery addr
wc.cell(7, 9, _clk("$I$3", 2))                 # I7  contact person
wc.cell(8, 9, _clk("$I$3", 3))                 # I8  mobile
wc.cell(9, 9, cidx("G"))                       # I9  TM no
wc.cell(10, 9, cidx("Q"))                      # I10 ordered qty
wc.cell(11, 9, cidx("M"))                      # I11 batch start
wc.cell(12, 9, cidx("N"))                      # I12 batch end
wc.cell(13, 9, cidx("F"))                      # I13 quantity
wc.cell(14, 9, cidx("E"))                      # I14 grade
wc.cell(15, 9, cidx("O"))                      # I15 slump (raw)
wc.cell(16, 9, cidx("P"))                      # I16 cube (raw)
wc.cell(17, 9, cidx("H"))                      # I17 driver
wc.cell(18, 9, f"=IF($C$4=\"\",\"\",IF($I$5=\"\",$I$13,SUMIFS({DBN}!$F$5:$F${TEND},{DBN}!$C$5:$C${TEND},$I$5)))")  # I18 cumulative

DATEFMT = "dd-mmm-yyyy"; TIMEFMT = "h:mm AM/PM"; NUMFMT = "#,##0.00"

def ch_pair(ws, r, llab, lref, rlab, rref, lfmt=None, rfmt=None):
    a = ws.cell(r, 2, llab); a.font = Font(bold=True, color="33425C", size=9)
    a.alignment = Alignment(indent=1, vertical="center"); a.border = B
    cc = ws.cell(r, 3, ":"); cc.alignment = Alignment(horizontal="center"); cc.border = B
    dv = ws.cell(r, 4, lref); dv.font = Font(bold=True, color=NAVY, size=10)
    dv.alignment = Alignment(indent=1, vertical="center"); dv.border = B
    if lfmt: dv.number_format = lfmt
    e = ws.cell(r, 5, rlab); e.font = Font(bold=True, color="33425C", size=9)
    e.alignment = Alignment(indent=1, vertical="center"); e.border = B
    fc = ws.cell(r, 6, ":"); fc.alignment = Alignment(horizontal="center"); fc.border = B
    gv = ws.cell(r, 7, rref); gv.font = Font(bold=True, color=NAVY, size=10)
    gv.alignment = Alignment(indent=1, vertical="center"); gv.border = B
    if rfmt: gv.number_format = rfmt
    ws.row_dimensions[r].height = 19

def challan_copy(ws, top, tag):
    def band(r, text, font, fill=None, h=None):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(r, 2, text); c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cc in range(2, 8):
            ws.cell(r, cc).border = B
            if fill: ws.cell(r, cc).fill = fill
        if h: ws.row_dimensions[r].height = h
    band(top + 0, PLANT_NAME, Font(bold=True, color=GOLD, size=15), FILL_NAVY, 26)
    band(top + 1, PLANT_ADDR, Font(color=NAVY, size=8), None, 16)
    band(top + 2, OFFICE_ADDR, Font(color=NAVY, size=8), None, 14)
    band(top + 3, REG_LINE, Font(color="33425C", size=8), None, 14)
    band(top + 4, "RMC DELIVERY CHALLAN   " + tag, Font(bold=True, color=NAVY, size=13), FILL_GOLD, 22)
    # section headers
    ws.merge_cells(start_row=top + 5, start_column=2, end_row=top + 5, end_column=4)
    s1 = ws.cell(top + 5, 2, "CHALLAN DETAILS"); s1.font = Font(bold=True, color=WHITE, size=10)
    s1.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=top + 5, start_column=5, end_row=top + 5, end_column=7)
    s2 = ws.cell(top + 5, 5, "CONCRETE DETAILS"); s2.font = Font(bold=True, color=WHITE, size=10)
    s2.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(2, 8):
        ws.cell(top + 5, cc).fill = FILL_NAVY2; ws.cell(top + 5, cc).border = B
    ws.row_dimensions[top + 5].height = 18
    d = top + 6
    ch_pair(ws, d + 0,  "Date", "=$I$1", "Batch Start", "=$I$11", DATEFMT, TIMEFMT)
    ch_pair(ws, d + 1,  "Challan No.", "=$C$4", "Batch End", "=$I$12", None, TIMEFMT)
    ch_pair(ws, d + 2,  "Batch No", "=$I$2", "Quantity (Cum)", "=$I$13", None, NUMFMT)
    ch_pair(ws, d + 3,  "Client Name", "=$I$3", "Grade", "=$I$14")
    ch_pair(ws, d + 4,  "GST details", "=$I$4", "Cement Type/Brand", f'="{CEMENT_BRAND}"')
    ch_pair(ws, d + 5,  "Delivery Address", "=$I$6", "Admixture Type", f'="{ADMIX_TYPE}"')
    ch_pair(ws, d + 6,  "Contact Person", "=$I$7", "Slump (mm)", f'=IF($I$15="","{DEFAULT_SLUMP}",$I$15)')
    ch_pair(ws, d + 7,  "Mobile", "=$I$8", "Cube Casting", f'=IF($I$16="","{DEFAULT_CUBE}",$I$16)')
    ch_pair(ws, d + 8,  "TM No.", "=$I$9", "Driver / Contact", "=$I$17")
    ch_pair(ws, d + 9,  "Unloading Start", "", "Unloading End", "")
    ch_pair(ws, d + 10, "Ordered Qty (Cum)", "=$I$10", "Cumulative Qty (Cum)", "=$I$18", NUMFMT, NUMFMT)
    t = d + 11
    for i, term in enumerate(CHALLAN_TERMS):
        ws.merge_cells(start_row=t + i, start_column=2, end_row=t + i, end_column=7)
        c = ws.cell(t + i, 2, term); c.font = Font(color="33425C", size=7)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        for cc in range(2, 8): ws.cell(t + i, cc).border = B
        ws.row_dimensions[t + i].height = 15
    sg = t + len(CHALLAN_TERMS)
    ws.merge_cells(start_row=sg, start_column=2, end_row=sg, end_column=4)
    ws.cell(sg, 2, "Received by (Site) / Signature").font = Font(color="33425C", size=9)
    ws.merge_cells(start_row=sg, start_column=5, end_row=sg, end_column=7)
    fr = ws.cell(sg, 5, "For, AAKRUTI INFRA RMC PLANT"); fr.font = Font(bold=True, color=NAVY, size=9)
    fr.alignment = Alignment(horizontal="right", indent=1)
    for cc in range(2, 8): ws.cell(sg, cc).border = B
    ws.row_dimensions[sg].height = 30
    band(sg + 1, WEB_LINE, Font(color=BLUE, size=8), None, 14)
    return sg + 1

end2 = challan_copy(wc, 6, "(CUSTOMER COPY)")
end2 = challan_copy(wc, end2 + 2, "(OFFICE COPY)")
# whatsapp + print
cbtn = end2 + 2
wa_msg = ("\"*AAKRUTI INFRA - Delivery Challan*%0AChallan No: \"&$C$4&\"%0ADate: \"&TEXT($I$1,\"dd-mmm-yyyy\")"
          "&\"%0AClient: \"&$I$3&\"%0AGrade: \"&$I$14&\"%0AQty: \"&$I$13&\" Cum%0ATM No: \"&$I$9"
          "&\"%0ADriver: \"&$I$17&\"%0ABatch No: \"&$I$2")
wa_url = ("=HYPERLINK(\"https://wa.me/\"&IF($I$8=\"\",\"\",\"91\"&$I$8)&\"?text=\"&"
          f"SUBSTITUTE({wa_msg},\" \",\"%20\"),\"\U0001F4F2  Send Challan on WhatsApp\")")
wa_button(wc, f"B{cbtn}:D{cbtn}", wa_url, "WhatsApp")
note_button(wc, f"E{cbtn}:G{cbtn}", "\U0001F5A8 Print / Download  \u00b7  press Ctrl+P  \u00b7  Save as PDF  (both copies fit one page)", GOLD_BTN)
fit_one_page(wc, f"A1:G{cbtn + 1}")

# ============================================================
# BATCH REPORT  (SCHWING-style, split per batch size / mixer capacity)
# ============================================================
wbr = wb.create_sheet("Batch Report")
wbr.sheet_view.showGridLines = False
set_widths(wbr, [2, 15, 9, 11, 11, 11, 11, 10, 10, 13])
page_form_title(wbr, COMPANY, "BATCH SHEET REPORT  \u2022  split per batch size", 10)
# selector (cidx keys off $C$4)
wbr.cell(4, 2, "Select Challan No:").font = Font(bold=True, color=NAVY, size=12)
g = wbr.cell(4, 3, FIRST_CHALLAN); g.fill = FILL_INPUT; g.border = B
g.font = Font(bold=True, color=NAVY, size=12); g.alignment = Alignment(horizontal="center")
wbr.merge_cells("C4:D4")
for cc in (3, 4): wbr.cell(4, cc).border = B
add_list_validation(wbr, "ChallanList", "C4")
hint2 = wbr.cell(4, 6, "Pick a challan + set Mixer Cap \u2192 production auto-splits into batches.")
hint2.font = Font(italic=True, color="5B6B86", size=10)
wbr.merge_cells("F4:J4")
# header grid (left value=col C, right value=col G)
box(wbr, 6, 2, "Plant", f'="{COMPANY}"', vmerge=2)
box(wbr, 6, 6, "Mixer Cap (m\u00b3)", 1.0, vmerge=2, vfmt="#,##0.00", value_input=True)
box(wbr, 7, 2, "Batch Number", f'=IFERROR(IF($C$4="","",IF({cidx_idx("L")}="",$C$4,{cidx_idx("L")})),"")', vmerge=2)
box(wbr, 7, 6, "Batch Size (m\u00b3)", "=IF($G$6=\"\",\"\",$G$6)", vmerge=2, vfmt="#,##0.00")
box(wbr, 8, 2, "Recipe / Grade", cidx("E"), vmerge=2)
box(wbr, 8, 6, "Production Qty (m\u00b3)", cidx("F"), vmerge=2, vfmt="#,##0.00")
box(wbr, 9, 2, "Batch Date", cidx("A"), vmerge=2, vfmt="dd-mmm-yyyy")
box(wbr, 9, 6, "No. of Batches", "=IF(OR($G$8=\"\",$G$6=\"\",$G$6=0),\"\",ROUNDUP($G$8/$G$6,0))", vmerge=2, vfmt="#,##0")
box(wbr, 10, 2, "Batch Start", cidx("M"), vmerge=2, vfmt="h:mm AM/PM")
box(wbr, 10, 6, "Customer", cidx("D"), vmerge=2)
box(wbr, 11, 2, "Batch End", cidx("N"), vmerge=2, vfmt="h:mm AM/PM")
box(wbr, 11, 6, "Order No", cidx("C"), vmerge=2)
box(wbr, 12, 2, "Truck No", cidx("G"), vmerge=2)
box(wbr, 12, 6, "Site", f"=IFERROR(IF($C$4=\"\",\"\",VLOOKUP({cidx_idx('C')},{ORDERS_TBL},3,0)),\"\")", vmerge=2)
box(wbr, 13, 2, "Truck Driver", cidx("H"), vmerge=2)
box(wbr, 13, 6, "W/C Ratio", f"=IFERROR(IF($C$8=\"\",\"\",VLOOKUP($C$8,{MIX_TBL},8,0)),\"\")", vmerge=2, vfmt="0.00")
for r in range(6, 14):
    for c in range(2, 9):
        if wbr.cell(r, c).border != B: wbr.cell(r, c).border = B
# --- per-batch material matrix ---
mhead = 15
heads2 = ["Batch", "Qty m\u00b3", "Cement", "Sand", "Agg 20", "Agg 10", "Water", "Admix", "Total"]
for i, h in enumerate(heads2):
    wbr.cell(mhead, 2 + i, h)
style_header(wbr, mhead, 10, fill=FILL_NAVY2)
wbr.cell(mhead, 1).fill = PatternFill("solid", fgColor=WHITE)
matcols = list(range(4, 10))   # D..I  (Mix Design column = c - 2)
r16 = mhead + 1                # Target / m3
wbr.cell(r16, 2, "Target / m\u00b3")
wbr.cell(r16, 3, 1).number_format = "#,##0.00"
for c in matcols:
    wbr.cell(r16, c, f"=IFERROR(IF($C$8=\"\",\"\",VLOOKUP($C$8,{MIX_TBL},{c - 2},0)),\"\")").number_format = "#,##0.0"
wbr.cell(r16, 10, f"=IF($C$8=\"\",\"\",SUM(D{r16}:I{r16}))").number_format = "#,##0.0"
r17 = mhead + 2                # Target / batch
wbr.cell(r17, 2, "Target / Batch")
wbr.cell(r17, 3, "=IF($G$7=\"\",\"\",$G$7)").number_format = "#,##0.00"
for c in matcols:
    L = get_column_letter(c)
    wbr.cell(r17, c, f"=IF({L}{r16}=\"\",\"\",{L}{r16}*$C${r17})").number_format = "#,##0.0"
wbr.cell(r17, 10, f"=IF($C$8=\"\",\"\",SUM(D{r17}:I{r17}))").number_format = "#,##0.0"
NB = 16
first = mhead + 3
for k in range(1, NB + 1):
    r = first + (k - 1)
    wbr.cell(r, 2, f"Batch {k}")
    wbr.cell(r, 3, f"=IF(OR($G$9=\"\",{k}>$G$9),\"\",IF({k}<$G$9,$G$7,$G$8-($G$9-1)*$G$7))").number_format = "#,##0.00"
    for c in matcols:
        L = get_column_letter(c)
        wbr.cell(r, c, f"=IF($C{r}=\"\",\"\",{L}${r16}*$C{r})").number_format = "#,##0.0"
    wbr.cell(r, 10, f"=IF($C{r}=\"\",\"\",SUM(D{r}:I{r}))").number_format = "#,##0.0"
tr = first + NB                # Total set weight
wbr.cell(tr, 2, "TOTAL SET WEIGHT")
wbr.cell(tr, 3, "=IF($C$8=\"\",\"\",$G$8)").number_format = "#,##0.00"
for c in matcols:
    L = get_column_letter(c)
    wbr.cell(tr, c, f"=IF($C$8=\"\",\"\",{L}{r16}*$G$8)").number_format = "#,##0.0"
wbr.cell(tr, 10, f"=IF($C$8=\"\",\"\",SUM(D{tr}:I{tr}))").number_format = "#,##0.0"
# styling for matrix body
for r in range(r16, tr + 1):
    for c in range(2, 11):
        cell = wbr.cell(r, c); cell.border = B
        if r == tr:
            cell.fill = GOLD_BTN; cell.font = Font(bold=True, color=NAVY, size=9)
        elif c == 2:
            cell.fill = PatternFill("solid", fgColor="F1F4F9"); cell.font = Font(bold=True, color=NAVY, size=9)
        else:
            cell.fill = FILL_CALC; cell.font = Font(color="1F3D6B", size=9)
# mass of total
mr = tr + 1
wbr.merge_cells(start_row=mr, start_column=2, end_row=mr, end_column=9)
ml = wbr.cell(mr, 2, "Mass of Total Set Weight (kg)")
ml.font = Font(bold=True, color=NAVY, size=10); ml.alignment = Alignment(horizontal="right", indent=1)
mv = wbr.cell(mr, 10, f"=IF($C$8=\"\",\"\",J{tr})"); mv.number_format = "#,##0.0"
mv.font = Font(bold=True, color=NAVY, size=10); mv.fill = FILL_GOLD
for c in range(2, 11): wbr.cell(mr, c).border = B
# note
nr = mr + 1
wbr.merge_cells(start_row=nr, start_column=2, end_row=nr, end_column=10)
nt = wbr.cell(nr, 2, "Set weights are Mix Design targets \u00d7 each batch's size. Up to 16 batches shown; the total always reflects full production qty. Operator records actual weighbridge values on the plant controller.")
nt.font = Font(italic=True, color="5B6B86", size=8)
nt.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
wbr.row_dimensions[nr].height = 22
# whatsapp + print
bbtn = nr + 2
br_msg = ("\"*AAKRUTI INFRA - Batch Sheet*%0ABatch No: \"&$C$7&\"%0ADate: \"&TEXT($C$9,\"dd-mmm-yyyy\")"
          "&\"%0AGrade: \"&$C$8&\"%0AQty: \"&$G$8&\" m3 in \"&$G$9&\" batches%0ACustomer: \"&$G$10"
          "&\"%0ATruck: \"&$C$12&\"%0ATotal Mass: \"&TEXT(J" + str(tr) + ",\"#,##0\")&\" kg\"")
br_url = ("=HYPERLINK(\"https://wa.me/?text=\"&"
          f"SUBSTITUTE({br_msg},\" \",\"%20\"),\"\U0001F4F2  Send Batch Sheet on WhatsApp\")")
wa_button(wbr, f"B{bbtn}:E{bbtn}", br_url, "WhatsApp")
note_button(wbr, f"F{bbtn}:J{bbtn}", "\U0001F5A8 Print / Download  \u00b7  press Ctrl+P  \u00b7  Save as PDF", GOLD_BTN)
fit_one_page(wbr, f"A1:J{bbtn + 1}")
wbr.page_setup.orientation = "landscape"

# ============================================================
# MONTHLY CONSUMPTION REPORT
# ============================================================
wmc = wb.create_sheet("Monthly Consumption")
set_widths(wmc, [3, 12, 14, 13, 14, 14, 13, 14, 13])
title_block(wmc, "Monthly Consumption Report", "Month-wise raw-material consumed across the plant. Set the year.", 9)
wmc.cell(4, 2, "Year:").font = Font(bold=True, color=GOLD, size=12)
ymc = wmc.cell(4, 3, 2026); ymc.font = Font(bold=True, color=NAVY, size=12)
ymc.fill = FILL_INPUT; ymc.border = B; ymc.alignment = Alignment(horizontal="center")
YMC = "$C$4"
mchead = 6
mc_cols = ["Month", "Cement (kg)", "Sand (kg)", "Agg 20mm (kg)", "Agg 10mm (kg)",
           "Water (ltr)", "Admixture (kg)", "Diesel (ltr)"]
for i, h in enumerate(mc_cols):
    wmc.cell(mchead, 2 + i, h)
style_header(wmc, mchead, len(mc_cols) + 1, fill=FILL_NAVY2)
wmc.cell(mchead, 1).fill = PatternFill("solid", fgColor=WHITE)
PRO = "'Batch Production'"
FUE = "'Fuel Log'"
def cons_month(val_rng, date_rng, m):
    return (f"=SUMIFS({val_rng},{date_rng},\">=\"&DATE({YMC},{m},1),"
            f"{date_rng},\"<\"&DATE({YMC},{m}+1,1))")
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for mi, mname in enumerate(months, start=1):
    rr = mchead + mi
    wmc.cell(rr, 2, mname).font = Font(bold=True, color=NAVY)
    wmc.cell(rr, 3, cons_month(f"{PRO}!$F$5:$F${TEND}", f"{PRO}!$A$5:$A${TEND}", mi))
    wmc.cell(rr, 4, cons_month(f"{PRO}!$G$5:$G${TEND}", f"{PRO}!$A$5:$A${TEND}", mi))
    wmc.cell(rr, 5, cons_month(f"{PRO}!$H$5:$H${TEND}", f"{PRO}!$A$5:$A${TEND}", mi))
    wmc.cell(rr, 6, cons_month(f"{PRO}!$I$5:$I${TEND}", f"{PRO}!$A$5:$A${TEND}", mi))
    wmc.cell(rr, 7, cons_month(f"{PRO}!$J$5:$J${TEND}", f"{PRO}!$A$5:$A${TEND}", mi))
    wmc.cell(rr, 8, cons_month(f"{PRO}!$K$5:$K${TEND}", f"{PRO}!$A$5:$A${TEND}", mi))
    wmc.cell(rr, 9, cons_month(f"{FUE}!$C$5:$C${TEND}", f"{FUE}!$A$5:$A${TEND}", mi))
    for cc in range(2, 10):
        cell = wmc.cell(rr, cc); cell.border = B
        cell.fill = FILL_CALC if cc > 2 else PatternFill("solid", fgColor=WHITE)
        cell.font = F_CALC if cc > 2 else Font(bold=True, color=NAVY)
        if cc > 2: cell.number_format = "#,##0"
mctot = mchead + 13
wmc.cell(mctot, 2, "TOTAL")
for cc in range(3, 10):
    L = get_column_letter(cc)
    c = wmc.cell(mctot, cc, f"=SUM({L}{mchead+1}:{L}{mchead+12})"); c.number_format = "#,##0"
for cc in range(2, 10):
    cell = wmc.cell(mctot, cc); cell.fill = GOLD_BTN; cell.font = Font(bold=True, color=NAVY); cell.border = B
mcchart = BarChart(); mcchart.type = "col"; mcchart.grouping = "stacked"; mcchart.overlap = 100
mcchart.title = "Monthly Material Consumption (kg)"; mcchart.height = 8; mcchart.width = 20
mcdata = Reference(wmc, min_col=3, min_row=mchead, max_col=6, max_row=mchead + 12)
mccats = Reference(wmc, min_col=2, min_row=mchead + 1, max_row=mchead + 12)
mcchart.add_data(mcdata, titles_from_data=True); mcchart.set_categories(mccats)
wmc.add_chart(mcchart, "B21")

# ============================================================
# RATE CARD  (grade-wise selling rates: market & in-house)
# ============================================================
wrc = wb.create_sheet("Rate Card")
set_widths(wrc, [3, 8, 16, 12, 18, 18])
title_block(wrc, "Rate Card  -  RMC Selling Rates",
            "Grade-wise market & in-house rates. Reference for quoting orders.", 6)
rc_heads = ["Sr.", "Grade", "Unit", "Market Rate (\u20b9/m\u00b3)", "In-House Rate (\u20b9/m\u00b3)"]
for i, h in enumerate(rc_heads):
    wrc.cell(HDR_ROW, 2 + i, h)
style_header(wrc, HDR_ROW, 6, fill=FILL_NAVY2)
wrc.cell(HDR_ROW, 1).fill = PatternFill("solid", fgColor=WHITE)
rate_card = [
    ("RMC M 10", "Per m\u00b3", 4580, 4450), ("RMC M 15", "Per m\u00b3", 4850, 4650),
    ("RMC M 20", "Per m\u00b3", 5350, 5200), ("RMC M 25", "Per m\u00b3", 5450, 5350),
    ("RMC M 30", "Per m\u00b3", 5650, 5450), ("RMC M 35", "Per m\u00b3", 5850, 5600),
    ("RMC M 40", "Per m\u00b3", 6400, 6230), ("DLC", "Per m\u00b3", 4400, 4000),
]
for i, (grade, unit, mkt, inh) in enumerate(rate_card):
    rr = DATA_START + i
    wrc.cell(rr, 2, i + 1)
    wrc.cell(rr, 3, grade).font = Font(bold=True, color=NAVY)
    wrc.cell(rr, 4, unit)
    m = wrc.cell(rr, 5, mkt); m.number_format = '\u20b9#,##0'
    h = wrc.cell(rr, 6, inh); h.number_format = '\u20b9#,##0'
    for cc in range(2, 7):
        cell = wrc.cell(rr, cc); cell.border = B
        cell.fill = FILL_CALC if cc >= 5 else PatternFill("solid", fgColor=WHITE)

# ============================================================
# DAILY EXPENSES
# ============================================================
ws_exp = build_sheet("Daily Expenses", "Daily Expenses  -  Cash & Plant Costs",
            "Log every plant expense. Month-to-date total shows top-right.",
            [{"h": "Date", "w": 14, "type": "date"}, {"h": "Paid For", "w": 34},
             {"h": "Amount (\u20b9)", "w": 14, "type": "money"}, {"h": "Paid By", "w": 18},
             {"h": "Mode", "w": 12}, {"h": "Category", "w": 18}, {"h": "Remarks", "w": 30}],
            TXN_ROWS, None, FILL_NAVY2)
add_list_validation(ws_exp, "PayModeList", f"E{DATA_START}:E{TEND}")
ws_exp.cell(2, 9, "THIS MONTH \u20b9").font = Font(bold=True, color=GOLD, size=11)
mtd = ws_exp.cell(3, 9, f"=SUMIFS($C${DATA_START}:$C${TEND},$A${DATA_START}:$A${TEND},"
                        f"\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),$A${DATA_START}:$A${TEND},"
                        f"\"<\"&DATE(YEAR(TODAY()),MONTH(TODAY())+1,1))")
mtd.number_format = '\u20b9#,##0'; mtd.font = Font(bold=True, color=NAVY, size=14)
mtd.fill = FILL_GOLD; mtd.border = B
ws_exp.column_dimensions["I"].width = 16
# --- import real April 2026 daily expenses ---
for _i, (_d, _pf, _amt) in enumerate(APRIL_EXP):
    _rr = DATA_START + _i
    if _rr > TEND:
        break
    ws_exp.cell(_rr, 1, _d)
    ws_exp.cell(_rr, 2, _pf)
    ws_exp.cell(_rr, 3, _amt)

# ============================================================
# STAFF ATTENDANCE
# ============================================================
ws_att = build_sheet("Staff Attendance", "Staff Attendance  -  Daily Muster",
            "Mark each day. Hours auto-calc from In/Out time.",
            [{"h": "Date", "w": 14, "type": "date"}, {"h": "Employee", "w": 22},
             {"h": "Status", "w": 13}, {"h": "In Time", "w": 12, "type": "input", "numfmt": "time"},
             {"h": "Out Time", "w": 12, "type": "input", "numfmt": "time"},
             {"h": "Hours", "w": 10, "type": "calc",
              "formula": "=IF(OR(D{r}=\"\",E{r}=\"\"),\"\",ROUND(MOD(E{r}-D{r},1)*24,2))"},
             {"h": "Trips", "w": 8, "type": "int"}, {"h": "Remarks", "w": 30}],
            TXN_ROWS, None, FILL_NAVY2)
add_list_validation(ws_att, "StaffList", f"B{DATA_START}:B{TEND}")
add_list_validation(ws_att, "AttStatusList", f"C{DATA_START}:C{TEND}")

# ============================================================
# TRIP & KM  (per-vehicle running + diesel efficiency)
# ============================================================
ws_trip = build_sheet("Trip & KM", "Trip & KM  -  Vehicle Running Log",
            "Per-vehicle km, trips & diesel. KM-run and efficiency auto-calc.",
            [{"h": "Date", "w": 14, "type": "date"}, {"h": "Vehicle No", "w": 16},
             {"h": "Driver", "w": 20, "type": "calc",
              "formula": "=IFERROR(IF(B{r}=\"\",\"\",VLOOKUP(B{r}," + VEH_TBL + ",4,0)),\"\")"},
             {"h": "Opening KM", "w": 13, "type": "num"}, {"h": "Closing KM", "w": 13, "type": "num"},
             {"h": "KM Run", "w": 11, "type": "calc",
              "formula": "=IF(OR(D{r}=\"\",E{r}=\"\"),\"\",E{r}-D{r})"},
             {"h": "Trips", "w": 8, "type": "int"}, {"h": "Diesel (L)", "w": 11, "type": "num"},
             {"h": "Eff (km/L)", "w": 11, "type": "calc",
              "formula": "=IF(OR(F{r}=\"\",H{r}=\"\",H{r}=0),\"\",ROUND(F{r}/H{r},2))"},
             {"h": "Remarks", "w": 28}],
            TXN_ROWS, None, FILL_NAVY2)
add_list_validation(ws_trip, "VehicleList", f"B{DATA_START}:B{TEND}")

# ============================================================
# LISTS (hidden)
# ============================================================
wl = wb.create_sheet("Lists")
wl["A1"] = "Vehicle Status"
for i, v in enumerate(["Active", "Maintenance", "Inactive"], start=2): wl[f"A{i}"] = v
wl["B1"] = "Dispatch Status"
for i, v in enumerate(["Scheduled", "Dispatched", "Delivered", "Cancelled"], start=2): wl[f"B{i}"] = v
wl["C1"] = "Payment Mode"
for i, v in enumerate(["Cash", "UPI", "Cheque", "NEFT", "RTGS"], start=2): wl[f"C{i}"] = v
wl["D1"] = "Attendance Status"
for i, v in enumerate(["Present", "Absent", "Half Day", "Leave", "Week Off"], start=2): wl[f"D{i}"] = v
wl["E1"] = "Staff"
for i, v in enumerate(["NEHA JAGTAP", "KRUSHNA S BADE", "YASH PATEL", "PREM KUMAR", "MANOJ KUMAR"], start=2):
    wl[f"E{i}"] = v
defname("AttStatusList", "'Lists'!$D$2:$D$6")
defname("StaffList", "'Lists'!$E$2:$E$51")
wl.sheet_state = "hidden"

# ---- tab order & colours ----
order = ["Instructions", "Dashboard", "Monthly Summary", "Monthly Consumption",
         "Setup - Clients", "Setup - Mix Design", "Setup - Vehicles", "Setup - Drivers", "Setup - Materials",
         "Rate Card",
         "Orders", "Dispatch (Challan)", "Challan", "Batch Production", "Batch Report",
         "Material Receipt", "Fuel Log", "Daily Expenses", "Trip & KM", "Staff Attendance",
         "Payments", "Cube Test Register", "Stock Register", "Lists"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
cmap = {"Instructions": GOLD, "Dashboard": GOLD, "Monthly Summary": GOLD, "Monthly Consumption": GOLD,
        "Setup - Clients": "5B6B86", "Setup - Mix Design": "5B6B86", "Setup - Vehicles": "5B6B86",
        "Setup - Drivers": "5B6B86", "Setup - Materials": "5B6B86", "Rate Card": GOLD,
        "Orders": NAVY2, "Dispatch (Challan)": "1E7F4F", "Challan": GOLD, "Batch Production": NAVY2,
        "Batch Report": GOLD, "Material Receipt": NAVY2, "Fuel Log": NAVY2,
        "Daily Expenses": NAVY2, "Trip & KM": NAVY2, "Staff Attendance": NAVY2, "Payments": "1E7F4F",
        "Cube Test Register": "1E5F8F", "Stock Register": "1E5F8F"}
for s in wb._sheets:
    if s.title in cmap: s.sheet_properties.tabColor = cmap[s.title]

wb.save(OUT)
print("Saved", OUT, "| TXN_ROWS", TXN_ROWS, "MASTER_ROWS", MASTER_ROWS)
print("Sheets:", [s.title for s in wb._sheets])
